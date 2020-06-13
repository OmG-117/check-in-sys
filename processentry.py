import smtplib
import time

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from fpdf import FPDF
from openpyxl import load_workbook

################################################################################

page_width = 210
page_height = 297
margin_size = 10

class PDF(FPDF):
    def add_field(self, label, value):
        initial_x = self.get_x()
        self.set_font('Arial', 'B', 12)
        label_width = self.get_string_width(label)
        self.cell(label_width, 10, label, 0, 0)
        self.set_font('Arial', '', 12)
        self.set_xy(self.get_x(), self.get_y() + 2)
        self.multi_cell(90 - label_width, 6, value, 0, 1)
        self.set_xy(self.get_x(), self.get_y() + 2)
        self.set_x(initial_x)

    def add_header(self, entry_num):
        self.set_font('Arial', 'B', 24)
        self.cell(190, 10, f'Entry #{entry_num}', 'B', 0, 'C')
        self.ln(20)
        self.set_font('Arial', '', 14)
        self.set_x(110)
        self.cell(90, 6, time.strftime('%A, %d %B %Y'), 0, 0, 'R')
        self.set_x(10)
        self.cell(90, 6, 'Hotel', 0, 1)
        self.cell(90, 6, 'Address 1', 0, 1)
        self.cell(90, 6, 'Address 2', 0, 1)

    def add_guest_info_box(self, inputs):
        self.set_font('Arial', 'B', 16)
        self.cell(90, 10, 'Guest info', 0, 1)
        y1 = self.get_y()
        self.add_field('Guest name: ', inputs['Full name'])
        self.add_field('Father\'s name: ', inputs['Father\'s name'])
        address = inputs['Address line 1'] + '\n'
        address += inputs['Address line 2']
        address += '\n' if inputs['Address line 2'] else ''
        address += inputs['City'] + ' - ' + inputs['Pincode'] + '\n'
        address += inputs['State'] + ', ' + inputs['Country']
        self.add_field('Address: ', address)
        self.add_field('Nationality: ', inputs['Nationality'])
        self.add_field('Phone number: ', inputs['Phone number'])
        self.add_field('Email ID: ', inputs['Email ID'])
        self.add_field('Arriving from: ', inputs['Arriving from'])
        self.add_field('Purpose of visit: ', inputs['Purpose of visit'])
        y2 = self.get_y()
        self.set_y(y1)
        self.cell(90, y2 - y1, '', 1, 1)

    def add_stay_info_box(self, inputs, y2):
        self.set_font('Arial', 'B', 16)
        self.cell(90, 10, 'Stay info', 0, 1)
        y1 = self.get_y()
        self.add_field('Check in date: ', time.strftime('%d/%m/%y'))
        self.add_field('Check in time: ', time.strftime('%I:%M %p'))
        self.ln(10)
        self.add_field('Expected stay: ', inputs['Expected stay'] + ' days')
        self.add_field('No. of persons: ', inputs['No. of persons'])
        self.ln(10)
        self.add_field('Check out date: ', '')
        self.add_field('Check out time: ', '')
        self.set_y(y1)
        self.cell(90, y2 - y1, '', 1, 1)

    def add_signature_box(self):
        self.set_font('Arial', 'B', 16)
        self.cell(90, 10, 'Guest signature', 0, 1)
        y1 = self.get_y()
        self.image('data/signature.png', h = 30)
        self.set_y(y1)
        self.cell(90, 30, '', 1, 1)

    def add_notes_box(self, inputs):
        self.set_font('Arial', 'B', 16)
        self.cell(90, 10, 'Notes', 0, 1)
        y1 = self.get_y()
        self.set_font('Arial', '', 12)
        self.multi_cell(90, 6, inputs['Notes'], 0, 1)
        self.set_y(y1)
        self.cell(90, 30, '', 1, 1)


def generate_pdf(inputs, entry_num):
    pdf = PDF()
    pdf.add_page()
    pdf.add_header(entry_num)

    pdf.ln(10)

    start_y = pdf.get_y()
    pdf.add_guest_info_box(inputs)
    end_y = pdf.get_y()

    pdf.set_xy(110, start_y)
    pdf.set_left_margin(110)
    pdf.add_stay_info_box(inputs, end_y)

    pdf.set_left_margin(10)
    pdf.set_x(10)
    pdf.ln(10)

    start_y = pdf.get_y()
    pdf.add_notes_box(inputs)

    pdf.set_xy(110, start_y)
    pdf.set_left_margin(110)
    pdf.add_signature_box()


    pdf.set_left_margin(10)
    pdf.set_x(10)
    pdf.ln(10)

    return pdf

################################################################################


def add_entry_to_xlsx(inputs, entry_num):
    filename = 'data/entries.xlsx'

    wb = load_workbook(filename)
    ws = wb.active

    entry = [int(entry_num)]
    entry.append(inputs['Full name'])
    entry.append(inputs['Father\'s name'])
    entry.append(inputs['Address line 1'])
    entry.append(inputs['Address line 2'])
    entry.append(inputs['City'])
    entry.append(inputs['Pincode'])
    entry.append(inputs['State'])
    entry.append(inputs['Country'])
    entry.append(inputs['Nationality'])
    entry.append(inputs['Phone number'])
    entry.append(inputs['Email ID'])
    entry.append(inputs['No. of persons'])
    entry.append(inputs['Expected stay'])
    entry.append(inputs['Arriving from'])
    entry.append(inputs['Purpose of visit'])

    ws.append(entry)
    wb.save(filename)

################################################################################

def email_file(filename, subject = ''):
    smtp_server = 'smtp.gmail.com'
    port = 587

    with open('credentials.txt', 'r') as file:
        credentials = []
        for line in file:
            credentials.append(line[ :-1])

    sender_email = ''
    receiver_email = ''

    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    message['Subject'] = subject

    with open(filename, 'rb') as file:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file.read())

    encoders.encode_base64(part)

    filename = filename.split('/').pop()
    part.add_header('Content-Disposition', 'attachment', filename = filename)

    message.attach(part)
    text = message.as_string()

    tries = 0
    while True:
        try:
            with smtplib.SMTP(smtp_server, port) as server:
                server.starttls()
                server.login(*credentials)
                server.sendmail(sender_email, receiver_email, text)
        except smtplib.SMTPException as error:
            if type(error) is smtplib.SMTPAuthenticationError or tries >= 1000:
                raise
            else:
                tries += 1
                time.sleep(60)
        else:
            break

################################################################################

if __name__ == '__main__':
    entry_num = time.strftime('%y%m') + '0001'
    inputs = {'Full name': 'Guestname Lastname',
              'Father\'s name': 'Fathersname Lastname',
              'Address line 1': 'Apt. XX, Streetname St.',
              'Address line 2': 'Areaname',
              'City': 'Cityname',
              'Pincode': 'XXXXXX',
              'State': 'Statename',
              'Country': 'Countryname',
              'Nationality': 'Nationality',
              'Phone number': 'XXXXXXXXXX',
              'Email ID': 'email@domain.com',
              'No. of persons': 'X',
              'Expected stay': 'X',
              'Arriving from': 'Originname',
              'Purpose of visit': 'Purpose',
              'Notes': 'Notes'}
    pdf = generate_pdf(inputs, entry_num)
    pdf.output('data/last.pdf', 'F')
    pdf.output(f'data/archive/{entry_num}.pdf', 'F')
    add_entry(inputs, entry_num)
    #email_entry(entry_num)
